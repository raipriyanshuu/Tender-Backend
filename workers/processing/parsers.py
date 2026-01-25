from __future__ import annotations

import csv
import os
from typing import Iterable

import PyPDF2
from docx import Document
from openpyxl import load_workbook

from workers.core.errors import ParseError, PermanentError
from workers.utils.filesystem import get_file_type

# OCR imports (optional - only if ENABLE_OCR=true)
_OCR_AVAILABLE = False
try:
    import pytesseract
    from pdf2image import convert_from_path
    from PIL import Image
    _OCR_AVAILABLE = True
except ImportError:
    pass

# GAEB imports (optional - only if GAEB_ENABLED=true)
_GAEB_AVAILABLE = False
try:
    from lxml import etree
    _GAEB_AVAILABLE = True
except ImportError:
    pass


def _ocr_pdf_page(image) -> str:
    """Extract text from a single PDF page image using OCR."""
    if not _OCR_AVAILABLE:
        return ""
    try:
        return pytesseract.image_to_string(image, lang="deu+eng")
    except Exception:  # noqa: BLE001
        return ""


def parse_pdf(file_path: str, enable_ocr: bool = False, ocr_max_pages: int = 50) -> str:
    """Parse PDF with optional OCR fallback for scanned documents."""
    try:
        with open(file_path, "rb") as handle:
            reader = PyPDF2.PdfReader(handle)
            text_content = "\n".join(page.extract_text() or "" for page in reader.pages)
            
            # Check if PDF is scanned (no/minimal text)
            if enable_ocr and _OCR_AVAILABLE and len(text_content.strip()) < 100:
                print(f"[Parser] PDF appears scanned ({len(text_content)} chars), running OCR...")
                
                # Limit pages for OCR to prevent runaway jobs
                num_pages = len(reader.pages)
                pages_to_ocr = min(num_pages, ocr_max_pages)
                
                try:
                    # Convert PDF to images and OCR
                    images = convert_from_path(file_path, dpi=300, first_page=1, last_page=pages_to_ocr)
                    ocr_chunks = [_ocr_pdf_page(img) for img in images]
                    ocr_text = "\n\n=== PAGE BREAK ===\n\n".join(ocr_chunks)
                    
                    if len(ocr_text.strip()) > 100:
                        print(f"[Parser] OCR successful: {len(ocr_text)} chars from {pages_to_ocr} pages")
                        return ocr_text
                    else:
                        print(f"[Parser] OCR failed to extract meaningful text")
                        return text_content  # Return original (even if empty)
                except Exception as ocr_exc:  # noqa: BLE001
                    print(f"[Parser] OCR failed: {ocr_exc}")
                    return text_content  # Fallback to original text
            
            return text_content
    except Exception as exc:  # noqa: BLE001 - return consistent error types
        raise ParseError(f"Failed to parse PDF: {file_path}") from exc


def parse_word(file_path: str) -> str:
    try:
        doc = Document(file_path)
        return "\n".join(paragraph.text for paragraph in doc.paragraphs)
    except Exception as exc:  # noqa: BLE001 - return consistent error types
        raise ParseError(f"Failed to parse Word document: {file_path}") from exc


def parse_excel(file_path: str) -> str:
    try:
        wb = load_workbook(filename=file_path, data_only=True)
        chunks: list[str] = []
        for sheet in wb.worksheets:
            chunks.append(f"## {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join("" if value is None else str(value) for value in row)
                if row_text.strip():
                    chunks.append(row_text)
        return "\n".join(chunks)
    except Exception as exc:  # noqa: BLE001 - return consistent error types
        raise ParseError(f"Failed to parse Excel file: {file_path}") from exc


def parse_csv(file_path: str) -> str:
    try:
        chunks: list[str] = []
        with open(file_path, "r", encoding="utf-8", errors="ignore") as handle:
            reader = csv.reader(handle)
            for row in reader:
                row_text = " | ".join(str(cell) for cell in row)
                if row_text.strip():
                    chunks.append(row_text)
        return "\n".join(chunks)
    except Exception as exc:  # noqa: BLE001 - return consistent error types
        raise ParseError(f"Failed to parse CSV file: {file_path}") from exc


def parse_text(file_path: str) -> str:
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as handle:
            return handle.read()
    except Exception as exc:  # noqa: BLE001 - return consistent error types
        raise ParseError(f"Failed to parse text file: {file_path}") from exc


def parse_gaeb(file_path: str) -> str:
    """Parse GAEB XML files (German tender exchange format) into normalized text."""
    if not _GAEB_AVAILABLE:
        raise ParseError("GAEB parsing requires lxml library. Install with: pip install lxml")
    
    try:
        tree = etree.parse(file_path)
        root = tree.getroot()
        
        # GAEB XML has multiple namespaces and versions (X83, X84, X89, etc.)
        # We'll extract key elements in a version-agnostic way
        
        chunks = []
        chunks.append("=== GAEB LEISTUNGSVERZEICHNIS ===\n")
        
        # Extract project info
        for award in root.xpath("//*[local-name()='Award']"):
            title = award.xpath(".//*[local-name()='BoQTitle']/text()")
            if title:
                chunks.append(f"Projekt: {title[0]}\n")
        
        # Extract Leistungsverzeichnis (Bill of Quantities)
        for boq in root.xpath("//*[local-name()='BoQ']"):
            boq_info = boq.xpath(".//*[local-name()='BoQInfo']")
            if boq_info:
                for info in boq_info:
                    name = info.xpath(".//*[local-name()='Name']/text()")
                    if name:
                        chunks.append(f"\n## {name[0]}")
        
        # Extract items/positions
        for item in root.xpath("//*[local-name()='Item']"):
            # Item number/OZ
            oz = item.xpath(".//*[local-name()='OZ']/text()")
            oz_text = oz[0] if oz else "N/A"
            
            # Description
            desc = item.xpath(".//*[local-name()='Description']//text()")
            desc_text = " ".join(desc) if desc else ""
            
            # Quantity
            qty = item.xpath(".//*[local-name()='Qty']/text()")
            qty_text = qty[0] if qty else ""
            
            # Unit
            unit = item.xpath(".//*[local-name()='QU']/text()")
            unit_text = unit[0] if unit else ""
            
            # Unit price (if exists)
            unit_price = item.xpath(".//*[local-name()='UP']/text()")
            price_text = f" | Preis: {unit_price[0]}" if unit_price else ""
            
            if desc_text:
                chunks.append(f"{oz_text} | {desc_text} | Menge: {qty_text} {unit_text}{price_text}")
        
        # Extract notes/remarks
        for note in root.xpath("//*[local-name()='Note'] | //*[local-name()='Remark']"):
            note_text = " ".join(note.xpath(".//text()"))
            if note_text.strip():
                chunks.append(f"\nHinweis: {note_text}")
        
        result = "\n".join(chunks)
        
        if not result.strip() or len(result) < 50:
            # If extraction yielded nothing meaningful, try raw text
            all_text = " ".join(root.xpath("//text()"))
            if len(all_text.strip()) > 50:
                return all_text
            raise ParseError(f"GAEB file appears empty or unreadable: {file_path}")
        
        return result
        
    except etree.XMLSyntaxError as exc:
        raise ParseError(f"Invalid GAEB XML format: {file_path}") from exc
    except Exception as exc:  # noqa: BLE001
        raise ParseError(f"Failed to parse GAEB file: {file_path}") from exc


def parse_file(
    file_path: str,
    enable_ocr: bool = False,
    ocr_max_pages: int = 50,
    temp_file_path: str | None = None,
) -> str:
    """
    Parse file based on type, with optional OCR for scanned PDFs.
    
    Args:
        file_path: Original file path (for type detection)
        enable_ocr: Enable OCR for scanned PDFs
        ocr_max_pages: Maximum pages to OCR
        temp_file_path: Temporary file path (if downloaded from R2)
        
    Returns:
        Extracted text content
    """
    # Use temp file path if provided, otherwise use original path
    actual_path = temp_file_path if temp_file_path else file_path
    
    file_type = get_file_type(file_path)
    
    if file_type == "pdf":
        return parse_pdf(actual_path, enable_ocr=enable_ocr, ocr_max_pages=ocr_max_pages)
    if file_type == "word":
        return parse_word(actual_path)
    if file_type == "excel":
        return parse_excel(actual_path)
    if file_type == "csv":
        return parse_csv(actual_path)
    if file_type == "text":
        return parse_text(actual_path)
    if file_type == "gaeb":
        return parse_gaeb(actual_path)
    
    raise PermanentError(f"Unsupported file type: {file_path}")
